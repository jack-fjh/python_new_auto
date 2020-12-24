from openpyxl import load_workbook
wb=load_workbook(filename=r'C:\tools\git\git-cangku\RF_demo\data\route.xlsx')
sheet_ranges=wb['Sheet2']
print(sheet_ranges['A1'].value)
print(sheet_ranges.cell(row=1, column=1).value)
min_row=wb
print(min_row)